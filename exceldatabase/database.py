import openpyxl
from openpyxl.utils import get_column_letter
import os
from typing import List, Dict, Any, Optional


class ExcelDatabase:
    def __init__(self, file_path: str, sheet_name: str = "Sheet1"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.data = self.load_data()

    def load_data(self) -> List[Dict[str, Any]]:
        if not os.path.exists(self.file_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            wb.save(self.file_path)
            return []
        
        wb = openpyxl.load_workbook(self.file_path)
        if self.sheet_name not in wb.sheetnames:
            wb.create_sheet(self.sheet_name)
            wb.save(self.file_path)
            return []
        
        ws = wb[self.sheet_name]
        data = list(ws.values)
        if not data:
            return []
        
        headers = data[0]
        return [dict(zip(headers, row)) for row in data[1:] if any(row)]

    def save_data(self):
        wb = openpyxl.load_workbook(self.file_path)
        if self.sheet_name not in wb.sheetnames:
            wb.create_sheet(self.sheet_name)
        ws = wb[self.sheet_name]
        
        if not self.data:
            wb.save(self.file_path)
            return
        
        headers = list(self.data[0].keys())
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)
        for row in self.data:
            ws.append([row.get(header, "") for header in headers])
        
        wb.save(self.file_path)

    def select(self, query: Dict[str, Any] = {}) -> Optional[List[Dict[str, Any]]]:
        result = [row for row in self.data if all(row.get(k) == v for k, v in query.items())]
        return result if result else None

    def get_column_value(self, search_column: str, search_value: Any, target_column: str) -> Optional[Any]:
        for row in self.data:
            if row.get(search_column) == search_value:
                return row.get(target_column)
        return None

    def insert(self, new_row: Dict[str, Any]):
        self.data.append(new_row)
        self.save_data()

    def update(self, query: Dict[str, Any], update_data: Dict[str, Any]):
        for row in self.data:
            if all(row.get(k) == v for k, v in query.items()):
                row.update(update_data)
        self.save_data()

    def delete(self, query: Dict[str, Any]):
        self.data = [row for row in self.data if not all(row.get(k) == v for k, v in query.items())]
        self.save_data()

    def add_sheet(self, sheet_name: str, initial_data: List[Dict[str, Any]] = []):
        wb = openpyxl.load_workbook(self.file_path)
        if sheet_name in wb.sheetnames:
            raise ValueError(f'Sheet "{sheet_name}" already exists.')
        ws = wb.create_sheet(sheet_name)
        if initial_data:
            ws.append(list(initial_data[0].keys()))
            for row in initial_data:
                ws.append(list(row.values()))
        wb.save(self.file_path)

    def is_sheet_exists(self, sheet_name: str) -> bool:
        wb = openpyxl.load_workbook(self.file_path)
        return sheet_name in wb.sheetnames

    def get_all_sheet_names(self) -> List[str]:
        wb = openpyxl.load_workbook(self.file_path)
        return wb.sheetnames

    def get_column_data_count(self, column_name: str) -> int:
        return sum(1 for row in self.data if row.get(column_name))

    def add_column(self, column_name: str, default_value: Any = None):
        for row in self.data:
            if column_name not in row:
                row[column_name] = default_value
        self.save_data()

    def remove_column(self, column_name: str):
        for row in self.data:
            if column_name in row:
                del row[column_name]
        self.save_data()