from typing import List, Dict, Any, Optional
from openpyxl import load_workbook

Row = Dict[str, Any]

class ExcelDatabase:
    def __init__(self, file_path: str, sheet_name: str = 'Sheet1'):
        """
        Construct Database
        :param file_path: The file path of the Excel file (e.g., './example.xlsx').
        :param sheet_name: The name of the sheet to use. Defaults to 'Sheet1'.
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.data: List[Row] = self._load_data()

    def _load_data(self) -> List[Row]:
        """Load data from the Excel sheet into a list of dicts."""
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' does not exist.")
        ws = wb[self.sheet_name]

        rows = list(ws.values)
        if not rows:
            return []
        headers = [str(h) for h in rows[0]]
        return [
            {headers[i]: cell for i, cell in enumerate(row)}
            for row in rows[1:]
        ]

    def _save_data(self) -> None:
        """Save the current in-memory data back to the Excel file."""
        wb = load_workbook(self.file_path)
        if self.sheet_name in wb.sheetnames:
            std = wb[self.sheet_name]
            wb.remove(std)
        ws = wb.create_sheet(self.sheet_name)
        if not self.data:
            wb.save(self.file_path)
            return

        headers = list(self.data[0].keys())
        ws.append(headers)
        for row in self.data:
            ws.append([row.get(col) for col in headers])

        wb.save(self.file_path)

    def select(self, query: Row = {}) -> Optional[List[Row]]:
        """
        Select rows matching all key-value pairs in `query`.
        :return: list of matched rows, or None if none match.
        """
        result = [
            row for row in self.data
            if all(row.get(k) == v for k, v in query.items())
        ]
        return result if result else None

    def get_column_value(self, search_column: str, search_value: Any, target_column: str) -> Any:
        """
        Find the first row where search_column == search_value and return row[target_column].
        :return: value or None if not found
        """
        for row in self.data:
            if row.get(search_column) == search_value:
                return row.get(target_column)
        return None

    def insert(self, new_row: Row) -> None:
        """Insert a new row and save."""
        self.data.append(new_row)
        self._save_data()

    def update(self, query: Row, update_data: Row) -> None:
        """Update all rows matching `query` by merging in `update_data`, then save."""
        updated = []
        for row in self.data:
            if all(row.get(k) == v for k, v in query.items()):
                merged = {**row, **update_data}
                updated.append(merged)
            else:
                updated.append(row)
        self.data = updated
        self._save_data()

    def delete(self, query: Row) -> None:
        """Delete all rows matching `query`, then save."""
        self.data = [
            row for row in self.data
            if not all(row.get(k) == v for k, v in query.items())
        ]
        self._save_data()

    def add_sheet(self, sheet_name: str, initial_data: List[Row] = []) -> None:
        """
        Add a new sheet to the Excel file.
        :raises ValueError: if sheet already exists.
        """
        wb = load_workbook(self.file_path)
        if sheet_name in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists.")
        ws = wb.create_sheet(sheet_name)
        if initial_data:
            headers = list(initial_data[0].keys())
            ws.append(headers)
            for row in initial_data:
                ws.append([row.get(col) for col in headers])
        wb.save(self.file_path)

    def is_sheet_exists(self, sheet_name: str) -> Optional[int]:
        """
        Check if a sheet exists.
        :return: 1 if exists, else None.
        """
        wb = load_workbook(self.file_path)
        return 1 if sheet_name in wb.sheetnames else None

    def get_all_sheet_names(self) -> List[str]:
        """Return a list of all sheet names in the Excel file."""
        wb = load_workbook(self.file_path)
        return wb.sheetnames

    def get_column_datas_number(self, column_name: str) -> int:
        """
        Count non-empty values in a given column.
        :return: count of rows where column_name is not None/empty-string.
        """
        return sum(
            1 for row in self.data
            if row.get(column_name) not in (None, '')
        )

    def add_column(self, column_name: str, default_value: Any = None) -> None:
        """
        Add a new column to the active sheet with a default value.
        """
        for row in self.data:
            if column_name not in row:
                row[column_name] = default_value
        self._save_data()

    def remove_column(self, column_name: str) -> None:
        """
        Remove a column from all rows, then save.
        """
        for row in self.data:
            if column_name in row:
                del row[column_name]
        self._save_data()
